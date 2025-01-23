import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# Caminhos das pastas de entrada e saída
PASTA_INPUT = "C:\\data\\input"
PASTA_OUTPUT = "C:\\data\\output"

# Dicionários de mapeamento
model_mapping = {
    "CODE01": "OPT1",
    "CODE02": "OPT2",
    "CODE03": "OPT3",
    "CODE04": "OPT4",
    "CODE05": "OPT5",
    "CODE06": "OPT6",
    "CODE07": "OPT7",
    "CODE08": "DEF1",
    "CODE09": "DEF2",
    "CODE10": "DEF3",
    "CODE11": "DEF4",
    "CODE12": "DEF5",
    "CODE13": "DEF6",
    "CODE14": "DEF7",
    "CODE15": "DEF8",
    "CODE16": "DEF9",
    "CODE17": "OPT8",
    "CODE18": "OPT9",
    "CODE19": None,
    "CODE20": "OPT10",
    "CODE21": "OPT11",
    "CODE22": "OPT12",
    "CODE23": "OPT13",
    "CODE24": "OPT14",
    "CODE25": "OPT15",
    "CODE26": "OPT16",
    "CODE27": "OPT17",
    "CODE28": "OPT18",
    "CODE29": "OPT19",
    "CODE30": "OPT20",
    "CODE31": "OPT21",
    "CODE32": "OPT22",
    "CODE33": "OPT23",
    "CODE34": "OPT24",
    "CODE35": "OPT25",
    "CODE36": "OPT26",
    "CODE37": "OPT27",
    "CODE38": "OPT28",
    "CODE39": "OPT29",
    "CODE40": "OPT30",
    "CODE41": "OPT31",
    "CODE42": "OPT32",
    "CODE43": "OPT33"
}

revenda_mapping = {
    "REVENDA_A": "LOC1",
    "REVENDA_B": "LOC2",
    "REVENDA_C": "LOC3",
    "REVENDA_D": "LOC4",
    "REVENDA_E": "LOC5",
    "REVENDA_F": "LOC6",
    "REVENDA_G": "LOC7",
    "REVENDA_H": "LOC8",
    "REVENDA_I": "LOC9",
    "REVENDA_J": "LOC10"
}

def formatar_tabela_excel(caminho_arquivo):
    """Formata a planilha como tabela Excel."""
    wb = load_workbook(caminho_arquivo)
    ws = wb.active
    ws.title = "STOCK"

    max_col = ws.max_column
    max_row = ws.max_row
    ultima_coluna = get_column_letter(max_col)
    tabela_ref = f"A1:{ultima_coluna}{max_row}"
    tabela = Table(displayName="Tabela1", ref=tabela_ref)

    estilo = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=True
    )
    tabela.tableStyleInfo = estilo
    ws.add_table(tabela)

    try:
        wb.save(caminho_arquivo)
        print(f"Tabela formatada no arquivo: {caminho_arquivo}")
    except Exception as e:
        print(f"Erro ao salvar a tabela: {e}")

def processar_dados():
    arquivos = [f for f in os.listdir(PASTA_INPUT) if (f.endswith('.xlsx') or f.endswith('.xls')) and not f.startswith('~$')]
    data_corte_as = pd.Timestamp("2024-09-01")
    data_corte_hs = pd.Timestamp("2024-09-20")

    for arquivo in arquivos:
        caminho_arquivo = os.path.join(PASTA_INPUT, arquivo)

        try:
            df = pd.read_excel(caminho_arquivo, sheet_name=0)
            df['Data_Compra'] = pd.to_datetime(df['Data_Compra'], errors='coerce', dayfirst=True)

            df['Options'] = df['Model_Code'].map(model_mapping)

            df.loc[df['Model_Code'] == 'CODE08', 'Options'] = df.apply(
                lambda row: 'DEF1' if row['Data_Compra'] <= data_corte_as else 'DEF2', axis=1
            )

            df.loc[df['Model_Code'] == 'CODE19', 'Options'] = df.apply(
                lambda row: 'DEF3' if row['Data_Compra'] <= data_corte_hs else 'DEF4', axis=1
            )

            df['Days'] = pd.to_numeric(df['Days'], errors='coerce')

            def atualizar_tempo_estoque(dias):
                if pd.isna(dias):
                    return None
                elif dias <= 30:
                    return '0-30'
                elif dias <= 60:
                    return '31-60'
                elif dias <= 90:
                    return '61-90'
                elif dias <= 120:
                    return '91-120'
                elif dias <= 150:
                    return '121-150'
                else:
                    return '151+'

            df['Stock_Time'] = df['Days'].apply(atualizar_tempo_estoque)
            df.rename(columns={'Stock_Time': 'TIME'}, inplace=True)
            df['Dealership_Name'] = df['Dealership_Name'].replace(revenda_mapping)

            df.dropna(how='all', inplace=True)
            df = df.iloc[:-1]

            caminho_saida = os.path.join(PASTA_OUTPUT, 'STOCK_REPORT.xlsx')
            df.to_excel(caminho_saida, index=False, engine='openpyxl')
            formatar_tabela_excel(caminho_saida)

            print(f"Arquivo processado: {caminho_saida}")

        except Exception as e:
            print(f"Erro ao processar {arquivo}: {e}")

if __name__ == "__main__":
    processar_dados()
